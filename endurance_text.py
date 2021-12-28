from numpy import number
import pandas as pd
import openpyxl
from Excel.excel_functions import append_df_to_excel
import re
import json
import subprocess
import os
import sys
import numpy as np

from Excel.excel_functions import excel_read_col_row

# pd.options.display.float_format = '{:,.4f}'.format

config_json = open('config.json', 'r')
config_json = json.load(config_json)

endurance_test_config = config_json['endurance_test']

sheet_name = 'Sheet'

print('-----------------------------')
print('\nRemeber to edit the config.json file\n')
print('-----------------------------')
print('-----------------------------')
print('\nRun options\n')
print('-----------------------------')
print('\n')
print(
    '1. Transfer data from text (.txt) file to excel (.xlsx) file. Select this option if you wish to update the config.json file according to a new excel file cell format.'
)
print(
    '2. Read text (.txt) files list and extract the rows and cols as configured in config.json and re-format to new excel files. Additional options could be set in the config.json.'
)
print(
    '3. Read excel (.xlsx) files list and extract the rows and cols as configured in config.json and re-format to new excel files. Only the data would be transferred, and the options would be constant for all the files.'
)

print('-----------------------------')
user_selection = input('Enter your choice: ')
print('\n')


def execute_powershell(command: str):
    subprocess.Popen(['powershell.exe', command])


def format_to_xlsx(file_path: str,
                   file_config,
                   file_to_write,
                   initial_col: number = 1):
    # f'{file_path.replace(".xlsx", "")}_transfer.xlsx'

    number_of_cycles = file_config['number_of_cycles']
    number_of_points = file_config['number_of_points']
    row_margin_buffer = file_config['row_margin_buffer']
    cols_to_read = file_config['cols_to_read']
    header_text = file_config['header_text']

    print(f"Formating excel file from path: {file_path}")

    for cycle_number in range(int(number_of_cycles)):

        start_row = file_config['start_row'] + (
            cycle_number * (number_of_points + row_margin_buffer + 1)
        ) if row_margin_buffer != None else file_config['start_row'] + (
            cycle_number * number_of_points)

        # print(f"Formatting row: {start_row}")

        df = pd.read_excel(
            file_path,
            sheet_name=sheet_name,
            usecols=cols_to_read,
        )
        voltage_polarization_data = df.iloc[start_row:start_row +
                                            number_of_points]

        col_to_write = initial_col + (cycle_number * 2)

        if header_text != None:
            header_df = pd.DataFrame(data=[header_text])

            append_df_to_excel(
                df=header_df,
                filename=file_to_write,
                sheet_name=sheet_name,
                startrow=1,
                startcol=col_to_write,
            )

        append_df_to_excel(
            df=voltage_polarization_data.astype('float'),
            filename=file_to_write,
            sheet_name=sheet_name,
            startrow=2,
            startcol=col_to_write,
        )


def transfer_single_txt_to_xlsx(file_path: str,
                                folder_directory_to_transfer: str):

    wb = openpyxl.Workbook()
    ws = wb.worksheets[0]

    file = open(file_path, 'r+')

    data = file.readlines()  # read all lines at once

    for row_index in range(len(data)):
        # This will return a line of string data
        row = data[row_index].split()

        for col_index in range(len(row)):
            row_data = re.sub('Â', '', row[col_index])
            ws.cell(row=row_index + 1, column=col_index + 1).value = row_data

    excel_file_path = f"{file_path.replace('txt', 'xlsx')}"

    try:
        os.makedirs(folder_directory_to_transfer)

    except FileExistsError:
        # directory already exists
        pass

    xlsx_file_name_match = re.search(r'(/\w*.xlsx)$', excel_file_path)

    if xlsx_file_name_match == None:
        print(
            'Invalid "file_path_to_read" argument in the config.json. Include the "/" to indicate file directories and include the ".txt" file extension.'
        )
        sys.exit()

    xlsx_file_name_index = xlsx_file_name_match.start()
    xlsx_file_name = excel_file_path[xlsx_file_name_index:]

    path_to_transfer = folder_directory_to_transfer + xlsx_file_name
    wb.save(path_to_transfer)

    file.close()

    print(f'Transferred data to {excel_file_path}')
    # print('Opening the file...')

    # # Open the new Excel file after data is written to it
    # execute_powershell(f'Invoke-Item \"{excel_file_path}\"')


def transfer_single_csv_to_xlsx(file_path_to_read: str,
                                folder_dir_to_write: str,
                                file_path_to_write: str):

    wb = openpyxl.Workbook()
    ws = wb.worksheets[0]

    file = open(file_path_to_read, 'r+')

    data = file.readlines()  # read all lines at once

    for row_index in range(len(data)):
        # This will return a line of string data
        row = data[row_index].split()

        for col_index in range(len(row)):
            row_data = row[col_index]

            if 'E' in row_data:
                split = row_data.split('E')
                value = split[0]
                exp = split[1]

                # Format the float exponent value (With E)
                try:
                    value = '{:,.2f}'.format(float(value))
                    row_data = f'{value}E{exp}'

                except:
                    pass

            ws.cell(row=row_index + 1, column=col_index + 1).value = row_data

    try:
        os.makedirs(folder_dir_to_write)

    except FileExistsError:
        # directory already exists
        pass

    excel_file_path = file_path_to_write

    wb.save(excel_file_path)
    file.close()

    # xlsx_file_name_match = re.search(r'(\.*.xlsx)$', excel_file_path)

    # if xlsx_file_name_match == None:
    #     print(
    #         'Invalid "file_path_to_read" argument in the config.json. Include the "/" to indicate file directories and include the ".csv" file extension.'
    #     )
    #     sys.exit()

    # xlsx_file_name_index = xlsx_file_name_match.start()
    # xlsx_file_name = excel_file_path[xlsx_file_name_index:]

    # print(f'Transferred data to {excel_file_path}')
    # print('Opening the file...')

    # # Open the new Excel file after data is written to it
    # execute_powershell(f'Invoke-Item \"{excel_file_path}\"')


def transfer_txt_to_xlsx():

    text_files_to_transfer = endurance_test_config[
        'option_1_txt_files_to_transfer_to_excel']

    for file in text_files_to_transfer:
        file_path = file['file_path']
        folder_directory_to_transfer = file['folder_directory_to_transfer']

        transfer_single_txt_to_xlsx(
            file_path=file_path,
            folder_directory_to_transfer=folder_directory_to_transfer)


def format_txt_files():
    files_to_format = endurance_test_config['option_2_files_to_transfer']

    files_to_format_names = list(files_to_format.keys())

    # Transfer the .txt file to .xlsx file (Text to excel)
    for file_to_format_name in files_to_format_names:
        print(
            f'\n------------------- Formatting {file_to_format_name} ---------------\n'
        )
        files = files_to_format[file_to_format_name]

        # Hard transfer each file; direct transfer line by line from .txt to .xlsx
        for file in files:
            if (file['format']) == 'break':
                continue

            file_path = file['file_path_to_read']
            folder_directory_to_transfer = file['folder_directory_to_transfer']
            transfer_single_txt_to_xlsx(
                file_path,
                folder_directory_to_transfer=folder_directory_to_transfer)

        # Transfer and extract each file
        for file_index, file in enumerate(files):
            file_path = file['file_path_to_read']
            folder_directory_to_transfer = file['folder_directory_to_transfer']
            folder_path_to_write = file['folder_path_to_write']

            initial_col = (file_index * 2) + 1
            file_path_xlsx_results = folder_path_to_write + '/' + file_to_format_name

            file_path_xlsx_data = f'{file_path.replace(".txt", ".xlsx")}'

            # Getting the directory to the .xlsx files where the .txt data is transferred to
            xlsx_file_name_match = re.search(r'(/\w*.xlsx)$',
                                             file_path_xlsx_data)

            if xlsx_file_name_match == None:
                print(
                    'Invalid "file_path_to_read" argument in the config.json. Include the "/" to indicate file directories and include the ".txt" file extension.'
                )
                sys.exit()

            xlsx_file_name = file_path_xlsx_data[xlsx_file_name_match.start():]
            path_to_read = folder_directory_to_transfer + xlsx_file_name

            try:
                os.makedirs(folder_path_to_write)

            except FileExistsError:
                # directory already exists
                pass

            format_to_xlsx(file_path=path_to_read,
                           file_config=file,
                           file_to_write=file_path_xlsx_results,
                           initial_col=initial_col)


def format_csv_to_excel():
    config = endurance_test_config['option_3_config']
    root_dirs = config.keys()

    for root_dir in root_dirs:

        root_dir_config = config[root_dir]
        settings = root_dir_config['settings']

        for setting in settings:

            file_type = setting['file_type']
            files = setting['files']
            relative_folder_path_to_read = setting[
                'relative_folder_path_to_read']
            folder_dir_to_write = setting["folder_dir_to_write"]
            excel_file_name_to_write = setting["excel_file_name_to_write"]
            relative_folder_path_transfer_for_csv_files = setting[
                'relative_folder_path_transfer_for_csv_files']
            cols_to_read = setting["cols_to_read"]
            rows_to_read = setting["rows_to_read"]
            start_row_to_write = setting["start_row_to_write"]

            hardcode_cols_to_write = setting["cols_to_write"]["hardcode"]

            if '.xlsx' not in excel_file_name_to_write:

                print(
                    "Provide a valid `excel_file_name_to_write` argument in config.json. Do ensure that the `.xlsx` extension is present."
                )
                sys.exit()

            for index, file_path in enumerate(files):

                if file_type == 'csv':
                    transfer_single_csv_to_xlsx(
                        file_path_to_read=
                        f'{root_dir}{relative_folder_path_to_read}\{file_path}',
                        folder_dir_to_write=
                        f'{root_dir}{relative_folder_path_transfer_for_csv_files}',
                        file_path_to_write=
                        f'{root_dir}{relative_folder_path_transfer_for_csv_files}\\{file_path.replace("csv", "xlsx")}',
                    )

                    excel_file_path_to_read = f'{root_dir}{relative_folder_path_transfer_for_csv_files}\{file_path}'.replace(
                        'csv', 'xlsx')
                    df = excel_read_col_row(excel_file=excel_file_path_to_read,
                                            rows_to_read=rows_to_read,
                                            cols_to_read=cols_to_read)

                    try:
                        os.makedirs(folder_dir_to_write)

                    except FileExistsError:
                        # directory already exists
                        pass

                    # Append dataframe to main excel file
                    append_df_to_excel(
                        df=df,
                        filename=
                        f'{folder_dir_to_write}\\{excel_file_name_to_write}',
                        startrow=start_row_to_write,
                        startcol=hardcode_cols_to_write[index])

                elif file_type == 'xlsx':
                    df = excel_read_col_row(excel_file=file_path,
                                            rows_to_read=rows_to_read,
                                            cols_to_read=cols_to_read,
                                            sheet_name="Sheet1")


if user_selection == "1":
    transfer_txt_to_xlsx()

elif user_selection == "2":
    format_txt_files()

elif user_selection == "3":
    format_csv_to_excel()
