from numpy import dtype
from numpy.lib.index_tricks import index_exp
import openpyxl
import pandas as pd
import numpy as np

from copy import copy
from pathlib import Path
from typing import Sequence, Union, Optional
from openpyxl import load_workbook


def excel_read_col_row(excel_file, rows_to_read: Sequence[int],
                       cols_to_read: Sequence[int]):

    df = pd.read_excel(
        excel_file,
        usecols=cols_to_read,
        header=None,
    )
    rows_range = range(rows_to_read[0] - 1, rows_to_read[1])

    return df.loc[rows_range]


def copy_excel_cell_range(
    src_ws: openpyxl.worksheet.worksheet.Worksheet,
    min_row: int = None,
    max_row: int = None,
    min_col: int = None,
    max_col: int = None,
    tgt_ws: openpyxl.worksheet.worksheet.Worksheet = None,
    tgt_min_row: int = 1,
    tgt_min_col: int = 1,
) -> openpyxl.worksheet.worksheet.Worksheet:

    if tgt_ws is None:
        tgt_ws = src_ws

    for row in src_ws.iter_rows(min_row=min_row,
                                max_row=max_row,
                                min_col=min_col,
                                max_col=max_col):
        for cell in row:
            tgt_ws.cell(
                row=cell.row + tgt_min_row - 1,
                column=cell.col_idx + tgt_min_col - 1,
                value=cell.value,
            )

    return tgt_ws


def append_df_to_excel(filename: Union[str, Path],
                       df: pd.DataFrame,
                       sheet_name: str = 'Sheet1',
                       startcol: Optional[int] = 0,
                       startrow: Optional[int] = 0,
                       header=False,
                       index=False,
                       **to_excel_kwargs) -> None:

    filename = Path(filename)
    file_exists = filename.is_file()

    if file_exists:
        wb = load_workbook(filename)
        sheet_names = wb.sheetnames
        sheet_exists = sheet_name in sheet_names
        sheets = {ws.title: ws for ws in wb.worksheets}

    with pd.ExcelWriter(
            filename.with_suffix(".xlsx"),
            engine="openpyxl",
            mode="a" if file_exists else "w",
            if_sheet_exists="new" if file_exists else None,
    ) as writer:
        if file_exists:
            # try to open an existing workbook
            writer.book = wb

            # copy existing sheets
            writer.sheets = sheets

        # write out the DataFrame to an ExcelWriter
        df.to_excel(writer,
                    sheet_name=sheet_name,
                    startcol=startcol,
                    startrow=startrow,
                    header=header,
                    index=index,
                    **to_excel_kwargs)

    if file_exists and sheet_exists:
        # move (append) rows from new worksheet to the `sheet_name` worksheet
        wb = load_workbook(filename)

        # retrieve generated worksheet name
        new_sheet_name = set(wb.sheetnames) - set(sheet_names)

        if new_sheet_name:
            new_sheet_name = list(new_sheet_name)[0]

        # copy rows written by `df.to_excel(...)` to
        copy_excel_cell_range(
            src_ws=wb[new_sheet_name],
            tgt_ws=wb[sheet_name],
        )

        # remove new (generated by Pandas) worksheet
        del wb[new_sheet_name]
        wb.save(filename)
        wb.close()
