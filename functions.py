import openpyxl
import os
import subprocess
import sys
import shutil
import re
import math
import pandas as pd
import numpy as np
import copy

from Excel.excel_functions import append_df_to_excel


def pretty_print(msg: str):
    print('\n#######################################################')
    print(f'\n{msg}\n')
    print('#######################################################\n')


def pretty_print_error_msg(err_msg: str):

    print('\n-------------------------------------')
    print(f'\n{err_msg}\n')
    print('-------------------------------------\n')


def transfer_file_to_new_folder(current_file_dir: str, target_dir: str,
                                target_file_name: str):

    try:
        os.makedirs(target_dir)

    except FileExistsError:
        pass

    try:
        shutil.copyfile(current_file_dir, target_dir + "/" + target_file_name)

    except:
        pass


def transfer_single_csv_to_xlsx(path_to_csv: str, folder_dir_to_write: str,
                                file_name_to_write: str):

    wb = openpyxl.Workbook()
    ws = wb.worksheets[0]

    try:
        file = open(path_to_csv, 'r+')

    except:
        print(f'Failed to open "{path_to_csv}", file path does not exist.\n')
        sys.exit()

    data = file.readlines()  # read all lines at once

    for row_index in range(len(data)):
        # This will return a line of string data
        row = data[row_index].split()

        for col_index in range(len(row)):
            row_data = row[col_index]

            if 'E' in row_data:
                split = row_data.split('E')
                value = split[0]
                exp = split[1]

                # Format the float exponent value (With E)
                try:
                    value = '{:,.2f}'.format(float(value))
                    row_data = f'{value}E{exp}'

                except:
                    pass

            row_data = row_data.replace(',', '')
            ws.cell(row=row_index + 1, column=col_index + 1).value = row_data

    try:
        os.makedirs(folder_dir_to_write)

    except FileExistsError:
        # directory already exists
        pass

    excel_file_path = f'{folder_dir_to_write}\\{file_name_to_write}'

    wb.save(excel_file_path)
    file.close()


def transfer_single_txt_to_xlsx(file_path: str,
                                folder_directory_to_transfer: str):

    wb = openpyxl.Workbook()
    ws = wb.worksheets[0]

    file = open(file_path, 'r+')

    data = file.readlines()  # read all lines at once

    for row_index in range(len(data)):
        # This will return a line of string data
        row = data[row_index].split()

        for col_index in range(len(row)):
            row_data = re.sub('Â', '', row[col_index])
            ws.cell(row=row_index + 1, column=col_index + 1).value = row_data

    excel_file_path = f"{file_path.replace('txt', 'xlsx')}"

    try:
        os.makedirs(folder_directory_to_transfer)

    except FileExistsError:
        # directory already exists
        pass

    xlsx_file_name_match = re.search(r'(/\w*.xlsx)$', excel_file_path)

    if xlsx_file_name_match == None:
        pretty_print_error_msg(
            'Invalid "file_path_to_read" argument in the config.json. Include the "/" to indicate file directories and include the ".txt" file extension.'
        )
        sys.exit()

    xlsx_file_name_index = xlsx_file_name_match.start()
    xlsx_file_name = excel_file_path[xlsx_file_name_index:]

    path_to_transfer = folder_directory_to_transfer + xlsx_file_name
    wb.save(path_to_transfer)

    file.close()


def create_folder(folder_dir: str):
    try:
        os.makedirs(folder_dir)

    except FileExistsError:
        pass


def create_file_and_append_df_to_xlsx(xlsx_folder_dir: str,
                                      xlsx_file_name: str, df: pd.DataFrame,
                                      startrow: int, startcol: int):
    try:
        df = df.astype('float')

    except:
        pass

    xlsx_file_path_to_write = f'{xlsx_folder_dir}{xlsx_file_name}'

    try:
        os.makedirs(xlsx_file_path_to_write)

    except FileExistsError:
        pass

    try:
        # Append dataframe to main excel file
        append_df_to_excel(df=df,
                           filename=xlsx_file_path_to_write,
                           startrow=startrow,
                           startcol=startcol)

    except:
        print(
            f'Failed to write to excel file. Ensure that the target file path "{xlsx_file_path_to_write}" is not running/open.\n'
        )
        sys.exit()


def execute_powershell(command: str):
    try:
        subprocess.check_output(['powershell.exe', command])

    except subprocess.CalledProcessError:
        raise Exception()


def execute_powershell_function(file_dir: str, fn_name: str, fn_args: list):
    powershell_cmd = f"&{fn_name} "

    for index, fn_arg in enumerate(fn_args):

        powershell_cmd = powershell_cmd + fn_arg + ' ' if index != len(
            fn_args) - 1 else powershell_cmd + fn_arg

    try:
        cmd = ["powershell.exe", f". \"{file_dir}\";", powershell_cmd]
        subprocess.check_output(cmd)

    except subprocess.CalledProcessError:
        raise Exception()


def order_files_according_to_config(files_to_order: list[str],
                                    ordered_values_config: list[str]):

    filtered_files_to_order = []
    ordered_value_fields = {}

    for config in ordered_values_config:
        ordered_value_fields[config] = []

    for file in files_to_order:
        for config in ordered_values_config:

            match = re.search(f'(^|\-|\_|\.){config}(\-|\_|\.|$)', file)

            if match:
                ordered_value_field_list = ordered_value_fields[config]
                ordered_value_field_list.append(file)

                ordered_value_fields[config] = ordered_value_field_list

                if file not in filtered_files_to_order:
                    filtered_files_to_order.append(file)

    ordered_files = copy.copy(filtered_files_to_order)
    file_expected_index = 0

    for file in filtered_files_to_order:

        for ordered_file_index, ordered_file in reversed(
                list(enumerate(ordered_files))):

            # If same file name
            if ordered_file == file:
                file_expected_index = ordered_file_index
                continue

            pos_status = None

            for field in ordered_value_fields:
                ordered_file_matches = ordered_value_fields[field]

                if ordered_file in ordered_file_matches and file in ordered_file_matches:
                    continue

                elif ordered_file not in ordered_file_matches and file not in ordered_file_matches:
                    continue

                elif ordered_file not in ordered_file_matches:
                    # Break this loop but continue in the outer loop
                    pos_status = '0'
                    file_expected_index = ordered_file_index

                    break

                elif file not in ordered_file_matches:

                    # Break this loop and the outer loop too
                    pos_status = '1'
                    file_expected_index = ordered_file_index + 1

                    break

            else:
                continue  # only executed if the inner loop did NOT break

            # Only executed if the inner loop DID break
            if pos_status == None:
                continue

            else:
                if pos_status == '1':
                    break

        prev_index = ordered_files.index(file)

        ordered_files.insert(file_expected_index, file)
        del ordered_files[prev_index]

    return ordered_files
